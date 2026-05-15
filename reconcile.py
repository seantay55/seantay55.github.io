"""
Invoice Reconciliation Script v3
==================================
Matches clinic invoices to business invoices by shared reference number.
Compares PRE-GST subtotals on both sides (business and clinic).
Deducts Facilitation Fees / Transportation from business subtotal before comparing.
Flags cash payment invoices (containing "Administrative and Support Fees").
Moves matched clinic invoices to /reconciled, unmatched to /unmatched.
Outputs a colour-coded Excel reconciliation report.

Folder structure expected:
  input/
    business/   ← e.g. #35891.pdf
    clinic/     ← e.g. (35891) 3232.pdf  or  (#35891) 3232.pdf
  output/
    reconciled/
    unmatched/
    reconciliation_report.xlsx
"""

import os
import re
import shutil
import glob
from pathlib import Path
from decimal import Decimal, InvalidOperation

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Configuration ────────────────────────────────────────────────────────────

CASH_PAYMENT_TRIGGER = "Administrative and Support Fees"
CURRENCY = "SGD"

# Line items to deduct from business subtotal before matching
DEDUCTIBLE_FEES = [
    "Facilitation Fees",
    "Facilitation Fee",
    "Transportation",
    "Transport",
]

INPUT_BUSINESS_DIR = "input/business"
INPUT_CLINIC_DIR   = "input/clinic"
OUTPUT_RECONCILED  = "output/reconciled"
OUTPUT_UNMATCHED   = "output/unmatched"
OUTPUT_REPORT      = "output/reconciliation_report.xlsx"


# ─── Reference Extraction ─────────────────────────────────────────────────────

def extract_ref_from_business(filename: str) -> str | None:
    """e.g. '#35891.pdf' or '#35891.pdf.pdf' → '35891'"""
    name = Path(filename).stem
    name = re.sub(r'\.pdf$', '', name, flags=re.IGNORECASE)  # handle double extension
    match = re.search(r'#?(\d+)', name)
    return match.group(1) if match else None


def extract_ref_from_clinic(filename: str) -> str | None:
    """e.g. '(35891) 3232.pdf' or '(#36231) 21981.pdf' → '35891' / '36231'"""
    name = Path(filename).stem
    match = re.search(r'\(#?(\d+)\)', name)
    return match.group(1) if match else None


# ─── PDF Reading ──────────────────────────────────────────────────────────────

def extract_text_from_pdf(filepath: str) -> str:
    text = ""
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print(f"  ⚠ Could not read {filepath}: {e}")
    return text


def extract_tables_from_pdf(filepath: str) -> list:
    tables = []
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)
    except Exception as e:
        print(f"  ⚠ Could not extract tables from {filepath}: {e}")
    return tables


# ─── Amount Extraction ────────────────────────────────────────────────────────

def parse_decimal(raw: str) -> Decimal | None:
    """Clean and parse a string into a Decimal. Returns None if invalid."""
    clean = raw.strip().replace(',', '').replace('SGD', '').replace('$', '').strip()
    try:
        val = Decimal(clean)
        return val if val > 0 else None
    except InvalidOperation:
        return None


def extract_subtotal(text: str) -> Decimal | None:
    """
    Extract the pre-GST subtotal from invoice text.

    Handles multiple invoice formats in priority order:

    1. Hospital final bill format:
       Sums 'Total Hospital Charges' + 'Total Doctor Charges' (both pre-GST).
       Doctor charges included even if zero.

    2. Standard subtotal label:
       'Sub-Total', 'Subtotal', 'Sub Total', 'Net Amount'

    3. Fallback:
       Derives subtotal = Grand Total - GST
    """

    # ── Format 1: Hospital bill ────────────────────────────────────
    hosp_charges = None
    for m in re.finditer(r'Total\s+Hospital\s+Charges\s+([0-9,]+\.\d{2})', text, re.IGNORECASE):
        preceding = text[max(0, m.start()-20):m.start()]
        if 'with' not in preceding.lower():
            hosp_charges = parse_decimal(m.group(1))

    doc_charges = Decimal("0")
    doc_match = re.search(r'TOTAL\s+DOCTOR\s+CHARGES\s+([0-9,]+\.\d{2})', text, re.IGNORECASE)
    if doc_match:
        val = parse_decimal(doc_match.group(1))
        if val:
            doc_charges = val

    if hosp_charges is not None:
        return hosp_charges + doc_charges

    # ── Format 2: Standard subtotal label ─────────────────────────
    subtotal_patterns = [
        r'Sub[-\s]?Total[\s:$]*([0-9,]+\.\d{2})',
        r'Subtotal[\s:$]*([0-9,]+\.\d{2})',
        r'Net\s+Amount[\s:$]*([0-9,]+\.\d{2})',
    ]
    for pattern in subtotal_patterns:
        matches = list(re.finditer(pattern, text, re.IGNORECASE))
        if matches:
            val = parse_decimal(matches[-1].group(1))
            if val:
                return val

    # ── Format 3: Fallback — Total minus GST ──────────────────────
    total = extract_total(text)
    gst   = extract_gst(text)
    if total and gst:
        return total - gst

    return None


def extract_total(text: str) -> Decimal | None:
    """Extract the grand total (including GST) from invoice text."""
    patterns = [
        r'(?:Grand\s+)?Total(?:\s+Due)?(?:\s+SGD)?[\s:$]*([0-9,]+\.\d{2})',
        r'Amount\s+Due[\s:$]*([0-9,]+\.\d{2})',
        r'Total\s+Amount[\s:$]*([0-9,]+\.\d{2})',
        r'Invoice\s+Total[\s:$]*([0-9,]+\.\d{2})',
        r'SGD\s*\$?\s*([0-9,]+\.\d{2})',
        r'\$\s*([0-9,]+\.\d{2})',
    ]
    found = []
    for pattern in patterns:
        for m in re.finditer(pattern, text, re.IGNORECASE):
            val = parse_decimal(m.group(1))
            if val:
                found.append(val)
    return found[-1] if found else None


def extract_gst(text: str) -> Decimal | None:
    """Extract the GST amount from invoice text."""
    patterns = [
        r'GST\s*(?:\(\s*\d+%\s*\)|\d+%)?\s*[\s:$]*([0-9,]+\.\d{2})',
        r'Tax\s*(?:\(\s*\d+%\s*\)|\d+%)?\s*[\s:$]*([0-9,]+\.\d{2})',
    ]
    for pattern in patterns:
        matches = list(re.finditer(pattern, text, re.IGNORECASE))
        if matches:
            val = parse_decimal(matches[-1].group(1))
            if val:
                return val
    return None


# ─── Deductible Fee Extraction ────────────────────────────────────────────────

def extract_deductible_fees(filepath: str, text: str) -> tuple[Decimal, list[dict]]:
    """
    Scan business invoice tables for deductible line items (Facilitation Fees,
    Transportation etc.). Returns (total_deduction, list of {label, amount}).
    Falls back to text scanning if table extraction finds nothing.
    """
    deductions     = []
    total_deduction = Decimal("0")

    # Method 1: table extraction
    tables = extract_tables_from_pdf(filepath)
    for table in tables:
        for row in table:
            if not row:
                continue
            desc = next((str(c).strip() for c in row if c and str(c).strip()), "")
            matched_label = next(
                (f for f in DEDUCTIBLE_FEES if f.lower() in desc.lower()), None
            )
            if not matched_label:
                continue
            for cell in reversed(row):
                if cell is None:
                    continue
                val = parse_decimal(str(cell))
                if val:
                    deductions.append({"label": desc, "amount": val})
                    total_deduction += val
                    break

    # Method 2: text fallback
    if not deductions:
        for line in text.split('\n'):
            matched_label = next(
                (f for f in DEDUCTIBLE_FEES if f.lower() in line.lower()), None
            )
            if not matched_label:
                continue
            amounts = re.findall(r'([0-9,]+\.\d{2})', line)
            if amounts:
                val = parse_decimal(amounts[-1])
                if val:
                    deductions.append({"label": matched_label, "amount": val})
                    total_deduction += val

    return total_deduction, deductions


# ─── Classification ───────────────────────────────────────────────────────────

def is_cash_payment_invoice(text: str) -> bool:
    return CASH_PAYMENT_TRIGGER.lower() in text.lower()


def ensure_dirs():
    for d in [INPUT_BUSINESS_DIR, INPUT_CLINIC_DIR, OUTPUT_RECONCILED, OUTPUT_UNMATCHED]:
        os.makedirs(d, exist_ok=True)


# ─── Main Reconciliation Logic ─────────────────────────────────────────────────

def run_reconciliation():
    ensure_dirs()

    print("\n" + "="*60)
    print("  INVOICE RECONCILIATION v3")
    print("="*60)

    # ── Step 1: Business invoices ──────────────────────────────────
    print("\n[1/4] Scanning business invoices...")

    business_invoices = {}
    cash_invoices     = []

    for filepath in glob.glob(os.path.join(INPUT_BUSINESS_DIR, "*.pdf")):
        filename = os.path.basename(filepath)
        ref      = extract_ref_from_business(filename)

        if not ref:
            print(f"  ⚠ Could not extract reference from: {filename} — skipping")
            continue

        text     = extract_text_from_pdf(filepath)
        is_cash  = is_cash_payment_invoice(text)
        subtotal = extract_subtotal(text)
        total    = extract_total(text)

        deduction_total, deductions = extract_deductible_fees(filepath, text)
        adjusted_subtotal = (subtotal - deduction_total) if subtotal else None

        info = {
            "ref":               ref,
            "filename":          filename,
            "filepath":          filepath,
            "total":             total,
            "subtotal":          subtotal,
            "deduction_total":   deduction_total,
            "deductions":        deductions,
            "adjusted_subtotal": adjusted_subtotal,
            "is_cash":           is_cash,
        }

        if is_cash:
            print(f"  💵 CASH PAYMENT  → {filename}  (ref: {ref})")
            cash_invoices.append(info)
        else:
            if deductions:
                labels = ", ".join(f"{d['label']} {CURRENCY} {d['amount']}" for d in deductions)
                print(f"  📄 Business      → {filename}  (ref: {ref}, subtotal: {CURRENCY} {subtotal}, deductions: {labels}, adjusted: {CURRENCY} {adjusted_subtotal})")
            else:
                print(f"  📄 Business      → {filename}  (ref: {ref}, subtotal: {CURRENCY} {subtotal})")
            business_invoices[ref] = info

    # ── Step 2: Clinic invoices ────────────────────────────────────
    print("\n[2/4] Scanning clinic invoices...")

    clinic_groups = {}

    for filepath in glob.glob(os.path.join(INPUT_CLINIC_DIR, "*.pdf")):
        filename = os.path.basename(filepath)
        ref      = extract_ref_from_clinic(filename)

        if not ref:
            print(f"  ⚠ Could not extract reference from: {filename} — skipping")
            continue

        text     = extract_text_from_pdf(filepath)
        subtotal = extract_subtotal(text)

        info = {
            "ref":      ref,
            "filename": filename,
            "filepath": filepath,
            "subtotal": subtotal,
        }

        clinic_groups.setdefault(ref, []).append(info)
        print(f"  🏥 Clinic        → {filename}  (ref: {ref}, subtotal: {CURRENCY} {subtotal})")

    # ── Step 3: Match & tally ─────────────────────────────────────
    print("\n[3/4] Matching & tallying...")

    report_rows = []
    all_refs    = set(business_invoices.keys()) | set(clinic_groups.keys())

    for ref in sorted(all_refs):
        biz     = business_invoices.get(ref)
        clinics = clinic_groups.get(ref, [])

        biz_subtotal       = biz["subtotal"]          if biz else None
        biz_total          = biz["total"]             if biz else None
        adjusted_subtotal  = biz["adjusted_subtotal"] if biz else None
        deductions         = biz["deductions"]        if biz else []
        deduction_total    = biz["deduction_total"]   if biz else Decimal("0")

        clinic_subtotal_sum = sum(
            c["subtotal"] for c in clinics if c["subtotal"] is not None
        ) if clinics else Decimal("0")

        clinic_files = [c["filename"] for c in clinics]

        # Compare adjusted business subtotal vs clinic subtotal sum
        compare_val = adjusted_subtotal if adjusted_subtotal is not None else biz_subtotal

        if not biz:
            status = "⚠ NO BUSINESS INVOICE"
        elif not clinics:
            status = "⚠ NO CLINIC INVOICES"
        elif biz_subtotal is None:
            status = "⚠ BUSINESS SUBTOTAL UNREADABLE"
        elif clinic_subtotal_sum == Decimal("0") and clinics:
            status = "⚠ CLINIC SUBTOTAL UNREADABLE"
        elif clinic_subtotal_sum == compare_val:
            status = "✅ MATCHED"
        else:
            status = "❌ MISMATCH"

        if deductions:
            labels = ", ".join(f"{d['label']} ({CURRENCY} {d['amount']})" for d in deductions)
            print(f"  Ref {ref}: Biz subtotal={biz_subtotal} | Deducted={deduction_total} ({labels}) | Adjusted={compare_val} | Clinic sum={clinic_subtotal_sum} | {status}")
        else:
            print(f"  Ref {ref}: Biz subtotal={biz_subtotal} | Clinic sum={clinic_subtotal_sum} | {status}")

        # Move clinic files
        for clinic in clinics:
            dest_dir  = OUTPUT_RECONCILED if status == "✅ MATCHED" else OUTPUT_UNMATCHED
            dest_path = os.path.join(dest_dir, clinic["filename"])
            shutil.copy2(clinic["filepath"], dest_path)

        deduction_notes = "\n".join(
            f"{d['label']}: {CURRENCY} {d['amount']}" for d in deductions
        ) if deductions else "—"

        report_rows.append({
            "ref":               ref,
            "business_file":     biz["filename"]        if biz else "—",
            "business_total":    float(biz_total)       if biz_total else None,
            "business_subtotal": float(biz_subtotal)    if biz_subtotal else None,
            "deduction_notes":   deduction_notes,
            "deduction_total":   float(deduction_total) if deduction_total else None,
            "adjusted_subtotal": float(adjusted_subtotal) if adjusted_subtotal else None,
            "clinic_files":      "\n".join(clinic_files) if clinic_files else "—",
            "clinic_subtotal":   float(clinic_subtotal_sum),
            "difference":        float(clinic_subtotal_sum - compare_val) if compare_val else None,
            "status":            status,
        })

    # Cash invoices → report only
    for c in cash_invoices:
        report_rows.append({
            "ref":               c["ref"],
            "business_file":     c["filename"],
            "business_total":    float(c["total"])    if c["total"]    else None,
            "business_subtotal": float(c["subtotal"]) if c["subtotal"] else None,
            "deduction_notes":   "—",
            "deduction_total":   None,
            "adjusted_subtotal": None,
            "clinic_files":      "—",
            "clinic_subtotal":   0.0,
            "difference":        None,
            "status":            "💵 CASH PAYMENT",
        })

    # ── Step 4: Report ────────────────────────────────────────────
    print("\n[4/4] Writing reconciliation report...")
    write_excel_report(report_rows)

    matched    = sum(1 for r in report_rows if "MATCHED"  in r["status"] and "CASH" not in r["status"])
    mismatched = sum(1 for r in report_rows if "MISMATCH" in r["status"])
    no_match   = sum(1 for r in report_rows if "NO "      in r["status"] or "UNREADABLE" in r["status"])
    adjusted   = sum(1 for r in report_rows if r["deduction_total"] and r["deduction_total"] > 0 and "CASH" not in r["status"])
    cash       = len(cash_invoices)

    print("\n" + "="*60)
    print(f"  ✅ Matched:              {matched}")
    print(f"  ✅ Matched w/ deductions:{adjusted}")
    print(f"  ❌ Mismatched:           {mismatched}")
    print(f"  ⚠  Unmatched/err:       {no_match}")
    print(f"  💵 Cash Payment:        {cash}")
    print(f"\n  Report saved to:  {OUTPUT_REPORT}")
    print(f"  Reconciled PDFs:  {OUTPUT_RECONCILED}/")
    print(f"  Unmatched PDFs:   {OUTPUT_UNMATCHED}/")
    print("="*60 + "\n")


# ─── Excel Report ─────────────────────────────────────────────────────────────

def write_excel_report(rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    ORANGE = PatternFill("solid", fgColor="FFEB9C")
    BLUE   = PatternFill("solid", fgColor="BDD7EE")
    TEAL   = PatternFill("solid", fgColor="D0EEE8")
    GREY   = PatternFill("solid", fgColor="D9D9D9")

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Reference #",
        "Business Invoice",
        f"Business Total incl. GST ({CURRENCY})",
        f"Business Subtotal ({CURRENCY})",
        "Deductions Applied",
        f"Total Deducted ({CURRENCY})",
        f"Adjusted Subtotal ({CURRENCY})",
        "Clinic Invoice(s)",
        f"Clinic Subtotal Sum ({CURRENCY})",
        f"Difference ({CURRENCY})",
        "Status",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2F4F7F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 40

    for i, row in enumerate(rows, start=2):
        status         = row["status"]
        has_deductions = row["deduction_total"] and row["deduction_total"] > 0

        if   "MATCHED"  in status and "CASH" not in status and has_deductions: fill = TEAL
        elif "MATCHED"  in status and "CASH" not in status:                    fill = GREEN
        elif "MISMATCH" in status:                                              fill = RED
        elif "CASH"     in status:                                              fill = BLUE
        elif "NO "      in status or "UNREADABLE" in status:                   fill = ORANGE
        else:                                                                   fill = GREY

        values = [
            row["ref"],
            row["business_file"],
            row["business_total"],
            row["business_subtotal"],
            row["deduction_notes"],
            row["deduction_total"],
            row["adjusted_subtotal"],
            row["clinic_files"],
            row["clinic_subtotal"],
            row["difference"],
            status,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in (3, 4, 6, 7, 9, 10) and val is not None:
                cell.number_format = f'"{CURRENCY}" #,##0.00'

        ws.row_dimensions[i].height = max(
            15 * max(row["clinic_files"].count("\n"), row["deduction_notes"].count("\n")) + 20, 20
        )

    widths = [14, 28, 26, 24, 35, 20, 24, 40, 24, 20, 26]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    matched    = sum(1 for r in rows if "MATCHED"  in r["status"] and "CASH" not in r["status"])
    mismatched = sum(1 for r in rows if "MISMATCH" in r["status"])
    no_match   = sum(1 for r in rows if "NO "      in r["status"] or "UNREADABLE" in r["status"])
    cash       = sum(1 for r in rows if "CASH"     in r["status"])
    adjusted   = sum(1 for r in rows if r["deduction_total"] and r["deduction_total"] > 0 and "CASH" not in r["status"])

    summary_data = [
        ("Invoice Reconciliation Summary", None),
        (None, None),
        ("Category",                      "Count"),
        ("✅ Matched",                     matched),
        ("✅ Matched with deductions",     adjusted),
        ("❌ Mismatched",                  mismatched),
        ("⚠ Unmatched / errors",          no_match),
        ("💵 Cash Payment",               cash),
        (None, None),
        ("Total processed",                len(rows)),
    ]

    for r_idx, (label, value) in enumerate(summary_data, 1):
        ws2.cell(row=r_idx, column=1, value=label)
        if value is not None:
            ws2.cell(row=r_idx, column=2, value=value)

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 12
    ws2["A1"].font = Font(bold=True, size=13)

    legend_start = len(summary_data) + 2
    ws2.cell(row=legend_start, column=1, value="Colour Legend").font = Font(bold=True)
    for j, (lfill, ldesc) in enumerate([
        (GREEN,  "✅ Matched"),
        (TEAL,   "✅ Matched with deductions"),
        (RED,    "❌ Mismatch"),
        (ORANGE, "⚠ Unmatched / error"),
        (BLUE,   "💵 Cash payment"),
    ], legend_start + 1):
        ws2.cell(row=j, column=1, value="").fill = lfill
        ws2.cell(row=j, column=2, value=ldesc).font = Font(size=9)

    wb.save(OUTPUT_REPORT)


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    run_reconciliation()
